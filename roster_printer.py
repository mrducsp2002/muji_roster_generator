from datetime import datetime, timedelta
import os
from helper import timespan_to_slot
from collections import defaultdict
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

def slot_to_time(slot):
    """Convert slot number back to time string (00:00 base, 15-min slots)."""
    # Start from midnight so slot 40 => 10:00 (40*15 = 600 minutes)
    start_time = datetime.strptime("00:00", "%H:%M")
    slot_time = start_time + timedelta(minutes=slot * 15)
    return slot_time.strftime("%H:%M")
def to_12h(time_str: str) -> str:
    """Convert 'HH:MM' (24h) to 'H:MM AM/PM' (12h)."""
    try:
        dt = datetime.strptime(time_str, "%H:%M")
        return dt.strftime("%I:%M %p").lstrip("0")
    except Exception:
        return time_str


def add_15_minutes(time_str):
    """Add 15 minutes to a time string - kept for potential future use"""
    time_obj = datetime.strptime(time_str, "%H:%M")
    new_time = time_obj + timedelta(minutes=15)
    return new_time.strftime("%H:%M")


def format_task_name(task):
    """Format task/department names for display"""
    task_names = {
        "FR": "Fitting Room",
        "GR": "Greeter",
        "R": "Register",
        "Break": "Break",
        "H": "H",
        "HH": "Home & Hardware",
        "L": "Ladies",
        "M": "Mens",
        "H&B": "Health & Beauty",
        "40": "Break (40min)",
        "10": "Break (10min)"
    }
    return task_names.get(task, task)


def print_roster_header(current_day, store_hours):
    """Print formatted header for the roster"""
    day_names = {
        "M": "Monday", "T": "Tuesday", "W": "Wednesday",
        "Th": "Thursday", "F": "Friday", "Sa": "Saturday", "Su": "Sunday"
    }

    print("=" * 80)
    print(f"STORE ROSTER - {day_names[current_day].upper()}")
    print(
        f"Store Hours: {store_hours[current_day][0]} - {store_hours[current_day][1]}")
    print("=" * 80)


def print_coverage_summary(roster, current_day, store_hours):
    """Print coverage summary for customer service tasks"""
    print("\n📋 CUSTOMER SERVICE COVERAGE SUMMARY")
    print("-" * 50)

    store_opening_slots = timespan_to_slot(store_hours[current_day])
    uncovered_slots = {"FR": [], "GR": [], "R": []}

    for slot in store_opening_slots:
        for task in ["FR", "GR", "R"]:
            if len(roster[slot][task]) == 0:
                uncovered_slots[task].append(slot)

    task_names = {"FR": "Fitting Room", "GR": "Greeter", "R": "Register"}

    for task, task_name in task_names.items():
        if uncovered_slots[task]:
            uncovered_slots_str = [
                f"Slot {slot}" for slot in uncovered_slots[task]]
            print(
                f"⚠️  {task_name}: UNCOVERED at {', '.join(uncovered_slots_str)}")
        else:
            print(f"✅ {task_name}: Fully covered")


def print_employee_schedule(roster, current_day, working_employees):
    """Print individual employee schedules"""
    print("\n👥 INDIVIDUAL EMPLOYEE SCHEDULES")
    print("-" * 50)

    # Get all employees and their shifts
    employee_schedules = defaultdict(list)

    for slot in roster:
        for task_or_dept in roster[slot]:
            for employee in roster[slot][task_or_dept]:
                employee_schedules[employee].append((slot, task_or_dept))

    # Sort employees by name
    for employee in sorted(employee_schedules.keys()):
        print(
            f"{employee} ({working_employees[employee]['department']} - Shift {working_employees[employee]['shift']}):")

        # Group consecutive same tasks
        schedule = employee_schedules[employee]
        if schedule:
            # Sort by slot number
            schedule.sort(key=lambda x: x[0])

            current_task = schedule[0][1]
            start_slot = schedule[0][0]
            end_slot = schedule[0][0]

            for i, (slot, task) in enumerate(schedule[1:], 1):
                if task == current_task and slot == end_slot + 1:
                    end_slot = slot
                else:
                    # Print previous task block
                    if start_slot == end_slot:
                        print(
                            f"  Slot {start_slot}: {format_task_name(current_task)}")
                    else:
                        print(
                            f"  Slots {start_slot}-{end_slot}: {format_task_name(current_task)}")
                    current_task = task
                    start_slot = slot
                    end_slot = slot

            # Print the last task block
            if start_slot == end_slot:
                print(f"  Slot {start_slot}: {format_task_name(current_task)}")
            else:
                print(
                    f"  Slots {start_slot}-{end_slot}: {format_task_name(current_task)}")
        else:
            print("  No assignments")


def print_hourly_breakdown(roster, current_day, store_hours):
    """Print slot-by-slot breakdown showing who's doing what"""
    print("\n🕒 SLOT-BY-SLOT BREAKDOWN")
    print("-" * 80)

    store_opening_slots = timespan_to_slot(store_hours[current_day])

    for slot in roster:
        if slot in store_opening_slots:
            print(f"Slot {slot}:")

            # Customer Service Tasks
            cs_tasks = []
            for task in ["FR", "GR", "R"]:
                if roster[slot][task]:
                    cs_tasks.append(f"{task}: {', '.join(roster[slot][task])}")
                else:
                    cs_tasks.append(f"{task}: EMPTY")

            print(f"  CS: {' | '.join(cs_tasks)}")

            # Other activities
            other_activities = []
            for activity in ["Break", "H"]:
                if roster[slot][activity]:
                    other_activities.append(
                        f"{activity}: {', '.join(roster[slot][activity])}")

            if other_activities:
                print(f"  Other: {' | '.join(other_activities)}")

            # Department work
            dept_work = []
            for dept in ["HH", "L's", "M's", "H&B"]:
                if roster[slot][dept]:
                    dept_work.append(
                        f"{dept}: {', '.join(roster[slot][dept])}")

            if dept_work:
                print(f"  Dept: {' | '.join(dept_work)}")

            print()  # Add space between slots


def print_register_coverage(roster, current_day, store_hours):
    """Print register staffing levels for each slot"""
    print("\n🧾 REGISTER STAFFING LEVELS")
    print("-" * 50)

    store_opening_slots = timespan_to_slot(store_hours[current_day])

    for slot in store_opening_slots:
        register_staff = roster[slot]["R"]
        staff_count = len(register_staff)

        if staff_count == 0:
            status = "⚠️  NO COVERAGE"
        elif staff_count == 1:
            status = "✅ MINIMAL"
        elif staff_count <= 3:
            status = "✅ GOOD"
        else:
            status = "🔥 HEAVY"

        time_str = slot_to_time(slot)
        print(f"Slot {slot} ({time_str}): {staff_count} people - {status}")
        if register_staff:
            print(f"  Staff: {', '.join(register_staff)}")
        print()


def print_statistics(roster, current_day, store_hours, working_employees):
    """Print useful statistics"""
    print("\n📊 ROSTER STATISTICS")
    print("-" * 40)

    store_opening_slots = timespan_to_slot(store_hours[current_day])

    # Count task assignments per employee
    employee_task_counts = defaultdict(lambda: defaultdict(int))
    total_cs_slots = 0

    for slot in store_opening_slots:
        for task in ["FR", "GR", "R"]:
            for employee in roster[slot][task]:
                employee_task_counts[employee][task] += 1
                total_cs_slots += 1

    print("Customer Service Task Distribution:")
    for employee in sorted(working_employees.keys()):
        fr_count = employee_task_counts[employee]["FR"]
        gr_count = employee_task_counts[employee]["GR"]
        r_count = employee_task_counts[employee]["R"]
        total = fr_count + gr_count + r_count

        if total > 0:
            print(
                f"  {employee}: FR({fr_count}) GR({gr_count}) R({r_count}) = {total} slots")

    print(f"\nTotal CS slots filled: {total_cs_slots}")
    print(f"Required CS slots: {len(store_opening_slots) * 3}")
    coverage_percentage = (
        total_cs_slots / (len(store_opening_slots) * 3)) * 100
    print(f"Coverage: {coverage_percentage:.1f}%")

    # Register staffing summary
    register_counts = []
    for slot in store_opening_slots:
        register_counts.append(len(roster[slot]["R"]))

    if register_counts:
        print(f"\nRegister Staffing Summary:")
        print(
            f"  Average staff per slot: {sum(register_counts) / len(register_counts):.1f}")
        print(f"  Minimum staff: {min(register_counts)}")
        print(f"  Maximum staff: {max(register_counts)}")

        # Count slots by staffing level
        no_coverage = sum(1 for count in register_counts if count == 0)
        minimal = sum(1 for count in register_counts if count == 1)
        good = sum(1 for count in register_counts if 2 <= count <= 3)
        heavy = sum(1 for count in register_counts if count > 3)

        print(f"  Slots with no coverage: {no_coverage}")
        print(f"  Slots with minimal coverage (1): {minimal}")
        print(f"  Slots with good coverage (2-3): {good}")
        print(f"  Slots with heavy coverage (4+): {heavy}")


def print_cs_coverage_totals(roster, current_day, store_hours):
    """Print total coverage count for each CS task per slot"""
    print("\n CS TASK COVERAGE TOTALS")
    print("-" * 50)

    store_opening_slots = timespan_to_slot(store_hours[current_day])

    print(f"{'Slot':<6} {'FR':<4} {'GR':<4} {'R':<4}")
    print("-" * 20)

    for slot in store_opening_slots:
        fr_count = len(roster[slot]["FR"])
        gr_count = len(roster[slot]["GR"])
        r_count = len(roster[slot]["R"])

        print(f"{slot:<6} {fr_count:<4} {gr_count:<4} {r_count:<4}")


def export_roster_to_excel(roster, current_day, working_employees, filename=None):
    output_dir = "roster_output"
    os.makedirs(output_dir, exist_ok=True)  # create folder if not exists

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"roster_{current_day}_{timestamp}.xlsx"

    # Ensure path points into roster_output
    filepath = os.path.join(output_dir, filename)

    # Prepare employee schedules with 1-hour buffers before opening and after closing
    original_slots = sorted(roster.keys())
    open_start_slot = min(original_slots) if original_slots else 0
    open_end_slot = max(original_slots) if original_slots else 0
    buffer_slots = 4  # 1 hour = 4 slots of 15 minutes
    start_with_buffer = max(0, open_start_slot - buffer_slots)
    end_with_buffer = open_end_slot + buffer_slots
    slots = list(range(start_with_buffer, end_with_buffer + 1))
    employee_schedules = {emp: {slot: "" for slot in slots}
                          for emp in working_employees.keys()}

    for slot in original_slots:
        for task_or_dept, employees in roster[slot].items():
            for emp in employees:
                if emp in employee_schedules:
                    task_name = task_or_dept
                    if employee_schedules[emp][slot]:
                        employee_schedules[emp][slot] += f" + {task_name}"
                    else:
                        employee_schedules[emp][slot] = task_name

    # Sort employees by shift then name
    sorted_employees = sorted(employee_schedules.keys(),
                              key=lambda emp: (working_employees.get(emp, {}).get("shift", 999), emp))
    sorted_employee_schedules = {
        emp: employee_schedules[emp] for emp in sorted_employees}

    df = pd.DataFrame.from_dict(sorted_employee_schedules, orient='index')

    # Column headers with slot info
    time_columns = {slot: f"{slot}" for slot in slots}
    df = df.rename(columns=time_columns)

    # Add First Name, Dept, Start, Finish columns
    first_names, dept_info, start_info, finish_info = [], [], [], []
    for emp in df.index:
        first_names.append(emp)
        if emp in working_employees:
            dept_info.append(working_employees[emp]["department"]) 
            start_info.append(to_12h(working_employees[emp]["shift"][0]))
            finish_info.append(to_12h(working_employees[emp]["shift"][1]))
        else:
            dept_info.append("")
            start_info.append("")
            finish_info.append("")
    df.insert(0, "First Name", first_names)
    df.insert(1, "Dept", dept_info)
    df.insert(2, "Start", start_info)
    df.insert(3, "Finish", finish_info)

    # Define colors for tasks and departments (RGB hex without '#')
    color_map = {
        # Departments
        "M's": "DAF2D0",        # Mens (incl. M's inner)
        "L's": "B5E6A2",        # Ladies (incl. W's inner)
        "K": "8ED973",          # Kids
        "M/K": "47D359",        # Mens + Kids
        "Acc.": "C1F0C8",       # Accessories
        "Fab": "DAE9F8",        # Fabric
        "Fab.": "DAE9F8",       # Fabric (alt label)
        "HW": "A6C9EC",         # Homewear
        "Stat.": "83CCEB",      # Stationery
        "H&B": "44B3E1",        # Health & Beauty
        "HH": "4D93D9",         # Household
        "F": "FBE2D5",          # Food
        # Tasks
        "40": "D9D9D9",         # 40min break
        "15": "D9D9D9",         # 15min break
        "10": "D9D9D9",         # 10min break
        "R": "BE5014",          # Register
        "FR": "F1A983",         # Fitting room
        "LD": "FFC000",         # Leader
        "GR": "F7C7AC",         # Greeter
        "H": "074F69"           # H
    }

    try:
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(
                writer, sheet_name=f'{current_day}_Schedule', index=False)
            worksheet = writer.sheets[f'{current_day}_Schedule']

            # Insert hour header row above the existing header to show grouped hours
            worksheet.insert_rows(1)
            first_time_col = 5  # A:First Name, B:Dept, C:Start, D:Finish, E: first slot
            # Populate minute row (row 2) as 00, 15, 30, 45 and hour row (row 1) with merged headers
            for i, slot in enumerate(slots):
                col_idx = first_time_col + i
                time_str = slot_to_time(slot)  # e.g. "10:15"
                hour_str, minute_str = time_str.split(":")
                # Set minute on row 2
                worksheet.cell(row=2, column=col_idx).value = minute_str
                # If this is the first quarter of the hour, write and merge the hour label across 4 columns
                if minute_str == "00":
                    worksheet.cell(row=1, column=col_idx).value = str(int(hour_str))
                    end_col = min(col_idx + 3, worksheet.max_column)
                    worksheet.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=end_col)
            # Center align header rows
            for i, slot in enumerate(slots):
                col_idx = first_time_col + i
                worksheet.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
                worksheet.cell(row=2, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
                # Header vertical borders matching body (thinner)
                is_hour_start = (slot % 4 == 0)
                left_side = Side(style="thin", color="000000") if is_hour_start else Side(style="hair", color="808080")
                for header_row in (1, 2):
                    hc = worksheet.cell(row=header_row, column=col_idx)
                    hb = hc.border
                    hc.border = Border(left=left_side, right=hb.right, top=hb.top, bottom=hb.bottom)

            # Set consistent column widths
            # A: First Name, B: Dept, C: Start, D: Finish
            worksheet.column_dimensions[get_column_letter(1)].width = 7
            worksheet.column_dimensions[get_column_letter(2)].width = 7
            worksheet.column_dimensions[get_column_letter(3)].width = 7
            worksheet.column_dimensions[get_column_letter(4)].width = 7
            for i, _ in enumerate(slots):
                col_idx = first_time_col + i
                worksheet.column_dimensions[get_column_letter(col_idx)].width = 3

            # Add vertical borders for time columns: thin black at hour start, hair grey at minutes
            thin_grey = Side(style="hair", color="808080")
            black_hour = Side(style="thin", color="000000")
            max_row = worksheet.max_row
            for i, slot in enumerate(slots):
                col_idx = first_time_col + i
                is_hour_start = (slot % 4 == 0)
                left_side = black_hour if is_hour_start else thin_grey
                for row_idx in range(1, max_row + 1):
                    c = worksheet.cell(row=row_idx, column=col_idx)
                    b = c.border
                    c.border = Border(
                        left=left_side,
                        right=b.right,
                        top=b.top,
                        bottom=b.bottom
                    )
            # Ensure a black right border at the end of the last time column
            last_time_col = first_time_col + len(slots) - 1
            for row_idx in range(1, max_row + 1):
                c = worksheet.cell(row=row_idx, column=last_time_col)
                b = c.border
                c.border = Border(
                    left=b.left,
                    right=black_hour,
                    top=b.top,
                    bottom=b.bottom
                )

            # Intentionally skip auto-fit here to preserve fixed widths (time=3, metadata=7)

            # Apply colors to all cells (including metadata and time slots)
            # Data now starts at row 3 because we inserted a header row
            for row in worksheet.iter_rows(min_row=3, min_col=1):
                for cell in row:
                    if cell.value is not None and str(cell.value).strip() != "":
                        # If multiple tasks, pick first for coloring
                        first_value = str(cell.value).split(" + ")[0]
                        fill_color = color_map.get(first_value, "FFFFFF")
                        cell.fill = PatternFill(start_color=fill_color,
                                                end_color=fill_color,
                                                fill_type="solid")
                    # Center align all cells
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    # Set font to Arial 8 for all cells
                    cell.font = Font(name="Arial", size=8)

            # Color the Dept column based on department
            for row_idx, emp in enumerate(list(df["First Name"]), start=3):
                cell = worksheet.cell(row=row_idx, column=2)
                dept = working_employees.get(emp, {}).get("department", "")
                # Normalize department keys to match color map
                if dept == "M":
                    dept = "M's"
                elif dept == "L":
                    dept = "L's"
                elif dept == "Acc":
                    dept = "Acc."
                elif dept == "Stat":
                    dept = "Stat."
                fill_color = color_map.get(dept, "FFFFFF")
                cell.fill = PatternFill(start_color=fill_color,
                                        end_color=fill_color,
                                        fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Arial", size=8)

            # Grey out pre-opening buffer; after closing buffer show dept if employee still working
            for row_idx, emp in enumerate(list(df["First Name"]), start=3):
                emp_info = working_employees.get(emp, {})
                emp_slots = []
                if emp_info and emp_info.get("shift"):
                    try:
                        emp_slots = list(timespan_to_slot(emp_info["shift"]))
                    except Exception:
                        emp_slots = []
                for i, slot in enumerate(slots):
                    col_idx = first_time_col + i
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if slot < open_start_slot:
                        # Pre-opening buffer always grey
                        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        cell.value = cell.value  # keep any value blank
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.font = Font(name="Arial", size=8)
                    elif slot > open_end_slot:
                        if slot in emp_slots:
                            # After-closing buffer: if still working, show department
                            dept = working_employees.get(emp, {}).get("department", "")
                            if dept == "M":
                                dept = "M's"
                            elif dept == "L":
                                dept = "L's"
                            elif dept == "Acc":
                                dept = "Acc."
                            elif dept == "Stat":
                                dept = "Stat."
                            cell.value = dept
                            fill_color = color_map.get(dept, "FFFFFF")
                            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        else:
                            # Not working: grey
                            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.font = Font(name="Arial", size=8)
                    else:
                        # Within store hours: if outside employee shift and empty, grey
                        if slot not in emp_slots and (cell.value is None or str(cell.value).strip() == ""):
                            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.font = Font(name="Arial", size=8)

            # Re-assert fixed widths at the very end to avoid any later overrides
            worksheet.column_dimensions[get_column_letter(1)].width = 12
            worksheet.column_dimensions[get_column_letter(2)].width = 7
            worksheet.column_dimensions[get_column_letter(3)].width = 7
            worksheet.column_dimensions[get_column_letter(4)].width = 7
            for i, _ in enumerate(slots):
                col_idx = first_time_col + i
                worksheet.column_dimensions[get_column_letter(col_idx)].width = 4

            # Create summary sheet using original store-open slots (exclude buffers)
            create_summary_sheet(writer, roster, current_day, original_slots)

        print(f"✅ Roster exported to: {filepath}")
        return filepath

    except Exception as e:
        print(f"❌ Error exporting to Excel: {str(e)}")
        return None

def create_summary_sheet(writer, roster, current_day, slots):
    """Create a summary sheet with coverage statistics"""

    summary_data = []

    # Coverage summary for each slot
    for slot in slots:
        slot_summary = {"Slot": slot, "Time": slot_to_time(slot)}

        # Count coverage for each task
        slot_summary["Fitting Room"] = len(roster[slot].get("FR", []))
        slot_summary["Greeter"] = len(roster[slot].get("GR", []))
        slot_summary["Register"] = len(roster[slot].get("R", []))
        slot_summary["On Break"] = len(roster[slot].get(
            "40", [])) + len(roster[slot].get("10", []))
        slot_summary["H"] = len(roster[slot].get("H", []))

        # Department coverage
        slot_summary["Home & Hardware"] = len(roster[slot].get("HH", []))
        slot_summary["Ladies"] = len(roster[slot].get("L's", []))
        slot_summary["Mens"] = len(roster[slot].get("M's", []))
        slot_summary["Health & Beauty"] = len(roster[slot].get("H&B", []))

        # Total employees working
        total_working = sum([
            len(employees) for task, employees in roster[slot].items()
            if task not in ["40", "10"]  # Exclude breaks from total
        ])
        slot_summary["Total Working"] = total_working

        summary_data.append(slot_summary)

    # Convert to DataFrame
    summary_df = pd.DataFrame(summary_data)

    # Write to Excel
    summary_df.to_excel(
        writer, sheet_name=f'{current_day}_Summary', index=False)

    # Format the summary sheet
    summary_worksheet = writer.sheets[f'{current_day}_Summary']

    # Auto-adjust column widths and center align all cells
    for column in summary_worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            cell.alignment = Alignment(horizontal="center", vertical="center")
        adjusted_width = max_length + 2
        summary_worksheet.column_dimensions[column_letter].width = adjusted_width
